import streamlit as st
import os
from utils.psi import preprocess_for_psi, calculate_psi
from utils.io_handler import save_to_excel
from utils.metrics import process_max_dpd_per_observation, deduplicate_gini, calculate_gini_metrics

st.set_page_config(page_title="Model Monitoring Tool", layout="wide")

# Judul Aplikasi
st.title("📊 Model Monitoring Tool")
st.markdown("""
Aplikasi ini menghitung:
- **PSI (Population Stability Index)**
- **AUROC, KS, dan Gini**
- Melakukan pencarian berdasarkan data DPD
""")

# Layout utama dibagi menjadi dua kolom
col1, col2 = st.columns(2)

# Kolom 1
with col1:
    st.header("1️⃣ Pilih Segmen, Periode, dan DPD")
    segment = st.selectbox("Pilih Segmen", ["SME", "Wholesale", "Mortgage"])

    dpd_options = [
        "0521", "0621", "0721", "0821", "0921", "1021", "1121", "1221",
        "0122", "0222", "0322", "0422", "0522", "0622", "0722", "0822", "0922", "1022", "1122", "1222",
        "0123", "0223", "0323", "0423", "0523", "0623", "0723", "0823", "0923", "1023", "1123", "1223",
        "0124", "0224", "0324", "0424", "0524", "0624", "0724", "0824", "0924", "1024", "1124", "1224",
        "0125", "0225", "0325", "0425", "0525", "0625", "0725", "0825", "0925", "1025", "1125", "1225",
        "0126", "0226", "0326", "0426", "0526", "0626", "0726", "0826", "0926", "1026", "1126", "1226"
    ]
    obs_options = [
        f"{year}.{str(month).zfill(2)}"
        for year in range(2021, 2026)
        for month in range(1, 13)
    ]

    selected_dpd = st.multiselect("🗂️ Pilih Search DPD", options=dpd_options, default=[])
    selected_obs = st.multiselect("📆 Pilih Periode Observasi", options=obs_options, default=[])

# Kolom 2
with col2:
    st.header("2️⃣ Input Direktori File")
    psi_file_path = st.text_input("Direktori untuk perhitungan PSI (file Excel): C:/SME/distribution dataset.xlsx ")
    gini_file_path = st.text_input("Direktori untuk Gini, KS, AUROC (file Excel): C:/SME/performance dataset.xlsx ")
    search_dpd_dir = st.text_input("Direktori untuk pencarian data DPD (folder): C:/input/Search DPD ")

# Proses monitoring
st.header("🚀 Proses Monitoring")

if st.button("Mulai Proses"):
    if psi_file_path and gini_file_path and search_dpd_dir and selected_dpd and selected_obs:
        try:
            st.info(f"🔍 Memproses data PSI untuk segmen: {segment} ...")

            # Proses PSI
            df_psi = preprocess_for_psi(psi_file_path, segment)
            psi_result = calculate_psi(df_psi, segment)

            # Proses Max DPD & Bad Flag
            result_dfs = process_max_dpd_per_observation(gini_file_path, search_dpd_dir, selected_obs, selected_dpd)

            # Proses Gini
            df_gini_dedup = deduplicate_gini(result_dfs)
            gini_df, gini_metrics_df, ks_value, auroc_value, gini_value = calculate_gini_metrics(df_gini_dedup, segment)

            # Simpan ke session state
            if segment == "Wholesale":
                # Untuk wholesale, simpan hasil terpisah untuk setiap size
                st.session_state["psi_results"] = psi_result
                st.session_state["segment"] = segment
            else:
                # Untuk segmen lain, simpan seperti biasa
                psi_value, psi_df = psi_result
                st.session_state["psi_df"] = psi_df
                st.session_state["psi_value"] = psi_value
                st.session_state["segment"] = segment
            
            st.session_state["max_dpd_all"] = result_dfs
            st.session_state["max_dpd_sheets"] = result_dfs
            st.session_state["gini_result"] = gini_metrics_df
            st.session_state["ks_value"] = ks_value
            st.session_state["auroc_value"] = auroc_value
            st.session_state["gini_value"] = gini_value

            # Tampilkan hasil PSI
            if segment == "Wholesale":
                st.success(f"✅ PSI berhasil dihitung untuk {segment}")
                for size_key, size_result in psi_result.items():
                    size_name = size_key.split("_")[1]  # Ambil "Large" atau "Medium"
                    psi_value = size_result["psi_value"]
                    psi_df = size_result["psi_df"]
                    st.markdown(f"### 📊 PSI untuk {size_name} Size: {psi_value:.4f}")
                    st.dataframe(psi_df)
            else:
                psi_value, psi_df = psi_result
                st.success(f"✅ PSI berhasil dihitung untuk {segment}: {psi_value:.4f}")
                st.dataframe(psi_df)

            # Tampilkan hasil Gini
            st.markdown("### 📈 Hasil Gini Metrics")
            st.dataframe(gini_metrics_df)
            st.markdown(f"""
            - **KS Value**: {ks_value * 100:.2f}%  
            - **AUROC**: {auroc_value * 100:.2f}%  
            - **Gini**: {gini_value * 100:.2f}%
            """)

        except Exception as e:
            st.error(f"❌ Gagal memproses data: {e}")
    else:
        st.warning("⚠️ Mohon isi semua field dan pilihan terlebih dahulu.")


# Simpan dan Unduh Hasil
if "psi_df" in st.session_state or "psi_results" in st.session_state:
    st.subheader("💾 Simpan dan Unduh Hasil")
    os.makedirs("output", exist_ok=True)

    # Simpan PSI
    if st.button("📥 Simpan Hasil PSI ke Excel"):
        if st.session_state['segment'] == "Wholesale":
            # Untuk wholesale, simpan setiap size ke sheet terpisah
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            output_psi = f"psi_{st.session_state['segment'].lower()}.xlsx"
            wb = Workbook()
            wb.remove(wb.active)
            
            for size_key, size_result in st.session_state["psi_results"].items():
                size_name = size_key.split("_")[1]  # Ambil "Large" atau "Medium"
                psi_df = size_result["psi_df"]
                # Bersihkan nilai NA sebelum disimpan ke Excel
                psi_df_clean = psi_df.fillna("").replace([float('inf'), float('-inf')], "")
                
                ws = wb.create_sheet(title=f"PSI_{size_name}")
                for r in dataframe_to_rows(psi_df_clean, index=False, header=True):
                    ws.append(r)
            
            wb.save(output_psi)
            st.success(f"✅ File PSI disimpan: `{output_psi}`")
        else:
            # Untuk segmen lain, simpan seperti biasa
            output_psi = f"psi_{st.session_state['segment'].lower()}.xlsx"
            save_to_excel(st.session_state["psi_df"], output_psi)
            st.success(f"✅ File PSI disimpan: `{output_psi}`")
        
        with open(output_psi, "rb") as f:
            st.download_button("⬇️ Unduh File PSI", data=f, file_name=os.path.basename(output_psi), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Simpan Max DPD dan Bad Flag
    if st.button("📥 Simpan Hasil Max DPD & Bad Flag"):
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        output_path = f"max_dpd_flag_{st.session_state['segment'].lower()}.xlsx"
        wb = Workbook()
        wb.remove(wb.active)

        # Tambah sheet untuk tiap observasi
        for sheet, df in st.session_state["max_dpd_sheets"].items():
            # Pastikan df adalah DataFrame yang valid
            if df is not None and hasattr(df, 'columns'):
                # Bersihkan nilai NA sebelum disimpan ke Excel
                df_clean = df.fillna("")  # Ganti NA dengan string kosong
                # Pastikan tidak ada nilai infinit atau NaN yang tersisa
                df_clean = df_clean.replace([float('inf'), float('-inf')], "")
                
                ws = wb.create_sheet(title=sheet[:31])  # Sheet name max 31 chars
                for r in dataframe_to_rows(df_clean, index=False, header=True):
                    ws.append(r)
            else:
                st.warning(f"⚠️ Data untuk sheet '{sheet}' tidak valid")

        wb.save(output_path)
        st.success(f"✅ File Max DPD disimpan: `{output_path}`")
        with open(output_path, "rb") as f:
            st.download_button("⬇️ Unduh File Max DPD", data=f, file_name=os.path.basename(output_path), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Simpan Gini
    if "gini_result" in st.session_state:
        output_gini = f"gini_{st.session_state['segment'].lower()}.xlsx"
        save_to_excel(st.session_state["gini_result"], output_gini)
        st.success(f"✅ File Gini disimpan: `{output_gini}`")
        with open(output_gini, "rb") as f:
            st.download_button("⬇️ Unduh File Gini", data=f, file_name=os.path.basename(output_gini), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
